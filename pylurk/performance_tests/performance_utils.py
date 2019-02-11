import os.path
from pylurk.core.conf import default_conf
from pylurk.core.lurk import LurkConf, ConfError, \
                             LurkUDPClient, LurkUDPServer, \
                             LurkTCPClient, LurkTCPServer, \
                             LurkHTTPClient, LurkHTTPServer
from pylurk.extensions.tls12 import LurkExt
from copy import deepcopy
from time import time, sleep
import openpyxl
import pandas as pd
from pylurk.utils.utils import start_server, stop_server, set_lurk, set_ssh
from pylurk.utils.data_plot import boxplot
from pylurk.utils.excel import write_to_excel
import multiprocessing as mp
import signal


def get_payload_params(mtype, **kwargs):
    '''
    This method will initiate a RequestPayload object based on the mtype and build the payload parameters accordingly given a set of parameters **kwargs
    :param mtype: request type [rsa_master, rsa_master_with_poh, rsa_extended_master, rsa_extended_master_with_poh, ecdhe]
    :param kwargs: parameters needed for building the payload
    :return: payload parameters to be used
    '''

    #get a RequestPayload object based on the mtype
    payload_request = LurkExt('client').get_ext_class()['request', mtype]

    #build payload based on a defined set of paramters
    payload_params = payload_request.build_payload(**kwargs)

    return payload_params

def latency_test (payload_params, connectivity_conf, graph_params, sheet_name, graph_path, excel_file = "results.xlsx",  thread = False, request_nb_list = [50], set_nb =20, remote_connection=False):
    '''
    This method is the main method that performs the latency tests based on the specified payload param. It will write:
    1- the test parameters in a sheet called  parameters_sheet_name = sheet_name+"_params"
    2- The latency results per set for each request_nb in latency_sheet_name = sheet_name+"_Val"
    3- Calculate the ratios of latencies based on reference column (ref in payload_params) and  save it in  ratio_sheet_name = sheet_name+"_Ratio"
    It will generate 2 box graphs based on the graph params, one for ratio and one for latency

    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform.
        each key in payload param should correspond to a key in connectivity conf.
        the other key, values in the dictionary of each item in payload params correspond to those desired for the test and should be similar to those defined in core/conf.py
        2 other keys, values are necessary:
            - column_name: name of the test as we want it to appear in the excel sheet
            - ref: corresponds to a column_name in the payload_params that we want to use to in the ration calculation (latency value of column_name/ latency value of ref)
    For example:
          payload_params = {
            'udp':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param connectivity_conf: dictionary with connectivity information of the client and the server identified by their keys.
        For each server configuration, we should have a client configuration with the same key. This key is also used in payload_params
         For example:
              connectivity_conf= {

            'udp':{
                'type': "udp",
                'ip_address': "127.0.0.1",
                'port': 6789,
                'key': join(data_dir, 'key_tls12_rsa_server.key'),
                'cert': join(data_dir, 'cert_tls12_rsa_server.crt'),
                'key_peer': join(data_dir, 'key_tls12_rsa_client.key'),
                'cert_peer': join(data_dir, 'cert_tls12_rsa_client.crt')
                 'remote_user':remote_user, #user name to set up ssh connection to remote server
                'password': server_password, #password of remote server
                'path_to_erilurk':"Desktop/HyameServer/projects/erilurk"
            }

    }
    :param graph_params a dictionary with all the information of the graph and data to plot (see utils/boxplot() for more info).
    :param sheet_name: name of the excel sheet where the data should be saved
    :param graph_path: path to the fig to save the graph (e.g., results/). The name of the graphs are ratio_sheet_name = sheet_name+"_Ratio".png , latency_sheet_name = sheet_name+"_Val".png.
            set to empty ("") if the fig is to be saved in the same directory
    :param excel_file:  path of excel file to save the results. The file is created if it does not exist
    :param thread: if set to true, it will enable the server to perform parallel execution of the requests using multi-threading
    :param request_nb_list: list of nb of requests  per set to report their latency. A client is generated for each request. The latency reported in for each request_nb in the list
    :param remote_connection: if set to true and remote_user is provided in the connectivity_conf, a connection to remote server will be initiated
    :param set_nb: nb of sets to test
    :return:
    '''
    #start by setting up a dictionnary of all parameters of the test to write in excel
    test_params = {
        'connectivity_conf': connectivity_conf,
        'payload_params' : payload_params,
        'set_nb':set_nb,
        'request_nb':request_nb_list,
        'thread':thread,
        'other_info': {'test_type': 'latency', #specify that we are testing the latency
                       'latency_unit': 'seconds', #specify that the reported latency results are in sec
                       }
    }

    parameters_sheet_name = sheet_name+"_params"
    latency_sheet_name = sheet_name+"_Val"
    ratio_sheet_name = sheet_name+"_Ratio"

    # print the parameters to the excel_file
    write_to_excel(excel_file, parameters_sheet_name, 1, 1, test_params = test_params)
    column_id = 1
    count =0
    print_set = True
    for request_nb in request_nb_list:
        if (count!=0):
            print_set=False
        #get latency by running the tests
        column_id = run_latency_test(payload_params, connectivity_conf, latency_sheet_name, excel_file = excel_file, thread = thread, request_nb= request_nb, set_nb=set_nb, column_id=column_id, print_set_id = print_set, remote_connection = remote_connection)
        count+=1
    #read the sheet containing the latencies and compute and write the ratios based on the reference results
    calculate_ratio(payload_params, latency_sheet_name, ratio_sheet_name, excel_file = excel_file)

    #create and save graph
    boxplot(ratio_sheet_name, excel_file, graph_params, graph_path+ratio_sheet_name+".png")
    boxplot(latency_sheet_name, excel_file, graph_params, graph_path+latency_sheet_name+".png")


def run_latency_test (payload_params, connectivity_conf, sheet_name, excel_file="results.xlsx", thread = False, request_nb=50, set_nb=20, column_id=1, print_set_id = True, remote_connection = False):
    '''
    This method performs the latency test and print the results into the specified sheet name
    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform.
        each key in payload param should correspond to a key in connectivity conf.
        the other key, values in the dictionary of each item in payload params correspond to those desired for the test and should be similar to those defined in core/conf.py
        2 other keys, values are necessary:
            - column_name: name of the test as we want it to appear in the excel sheet
            - ref: corresponds to a column_name in the payload_params that we want to use to in the ration calculation (latency value of column_name/ latency value of ref)
    For example:
          payload_params = {
            'udp':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param connectivity_conf: dictionary with connectivity information of the client and the server identified by their keys.
        For each server configuration, we should have a client configuration with the same key. This key is also used in payload_params
         For example:
              connectivity_conf= {
              'udp':{
                'type': "udp",
                'ip_address': "127.0.0.1",
                'port': 6789,
                'key': join(data_dir, 'key_tls12_rsa_server.key'),
                'cert': join(data_dir, 'cert_tls12_rsa_server.crt'),
                'key_peer': join(data_dir, 'key_tls12_rsa_client.key'),
                'cert_peer': join(data_dir, 'cert_tls12_rsa_client.crt')
                 'remote_user':remote_user, #user name to set up ssh connection to remote server
                'password': server_password, #password of remote server
                'path_to_erilurk':"Desktop/HyameServer/projects/erilurk"
            }

    }
    :param sheet_name: name of the excel sheet where the data should be saved
    :param excel_file:  path of excel file to save the results. The file is created if it does not exist
    :param thread:if set to true, it will allow the server to execute the clients requests in parallel using multithreading
    :param request_nb: nb of requests  per set to report their latency. it corresponds to the number of clients to create and each client generates one request
    :param set_nb: nb of sets to test
    :param column_id: column id in which we want to start writing in the excel sheet
    :param print_set_id: if set to true will print the set id in the excel sheet. This is useful when we have multiple call for this method and we want to print the set id once
    :param remote_connection: if set to true, and remote_user is provided in the connectivity_conf, a connection to remote server will be initiated
    :return: column_id in the excel sheet where we can start writing without overwriting any existing values

    '''

    row_id = 1


    # check if excel file exists and create it otherwise
    if (os.path.isfile(excel_file) == False):
        book = openpyxl.Workbook()
        book.save(excel_file)

    book = openpyxl.load_workbook(excel_file)

    # create the sheet if it does not exist in the mentioned excel_file
    if (sheet_name not in book.sheetnames):
        book.create_sheet(sheet_name)
        book.save(excel_file)

    # get the sheet
    sheet = book[sheet_name]
    test_count = 0

    if (print_set_id):
        sheet.cell(row=row_id, column=column_id).value = 'Set'
        test_count = 1

    for connectivity_key, test_params in payload_params.items():

        #enable remote connection only if remote_user is set
        if (remote_connection and 'remote_user' in connectivity_conf[connectivity_key].keys()):
            remote = True
        elif (remote_connection):
            print("Remote user not provided to enable remote connection --- running locally")
            remote = False
        else:
            remote = False

        #start server corresponding to the key
        process_id = start_server(connectivity_conf=connectivity_conf[connectivity_key], thread=thread, remote_connection=remote)

        # create a client
        client = set_lurk('client', connectivity_conf=connectivity_conf[connectivity_key], resolver_mode='stub')

        for params_value in test_params:

            # write the column name reflecting the test we are performing
            sheet.cell(row=row_id, column=column_id + test_count).value = params_value['column_name']

            # generate payload params
            generated_payload_params = get_payload_params(params_value['type'], args=params_value)

            # loop over the set nb
            for j in range(0, set_nb):

                time_start = time()

                for i in range(0, request_nb):

                    client.resolve([{'designation': 'tls12', \
                                     'version': 'v1', 'status': "request", \
                                     'type': params_value['type'], 'payload': generated_payload_params}])
                # get total processing time for all requests
                time_stop = time()

                # calculate the processing time for requestsNb requests and return it
                total_time = time_stop - time_start

                # write the set id
                sheet.cell(row=row_id + j + 1, column=column_id).value = j
                # write the total time to send and receive request_nb requests
                sheet.cell(row=row_id + j + 1, column=column_id + test_count).value = total_time  # to change columns

                # save the file at each set test to keep track of the test
                book.save(excel_file)

                # print some log info
                print(" %s resolutions in %s sec." % ( params_value['column_name'], total_time))

            test_count = test_count + 1

       #stop the server
        if(remote==True):
            stop_server(process_id, remote_host=connectivity_conf[connectivity_key]['ip_address'], remote_user=connectivity_conf[connectivity_key]['remote_user'], password = connectivity_conf[connectivity_key]['password'])
        else:
            stop_server(process_id)

    book.close()

    # column id succeeding the last column where some data was written
    return column_id+test_count


def calculate_ratio (payload_params, values_sheet_name, sheet_name, excel_file = "results.xlsx"):
    '''
    This method calculates the ratio values based on the average of the reference values mentioned in payload_params and save them in the mentioned sheet_name
    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform. For example:
          payload_params = {'rsa_master_ref_prf_sha256_pfs_sha256':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param values_sheet_name: the sheet name in excel_file containing the values to which we want to calculate the ratio
    :param sheet_name: sheet_name to save the calculates ratio. This sheet is created if it does not exists
    :param excel_file: path to the excel file containing the values to calculate the ratio and to save the latter
    '''
    try:

        book = openpyxl.load_workbook(excel_file)

        # create the results sheet if it does not exist in the mentioned excel_file
        if (sheet_name not in book.sheetnames):
            book.create_sheet(sheet_name)
            book.save(excel_file)


        # get the sheet where we want to store the ratio values
        ratio_sheet = book[sheet_name]

        row_id =1
        column_id =1
        test_count =1

        ratio_sheet.cell(row=row_id, column=column_id).value = 'Set'

        #read the sheet containing the values to which we want to calculate the ratio
        df = pd.read_excel(excel_file, sheet_name=values_sheet_name, engine =None)

        for connectivity_key, test_params in payload_params.items():
            # loop over the different tests (columns)
            for params_value in test_params:

                # write the column name reflecting the test
                ratio_sheet.cell(row=row_id, column=column_id + test_count).value = params_value['column_name']

                #loop over the latency values of each column
                for i in range (0, len(df[params_value['column_name']])):

                    #calculate the ratio based on the average latency of the ref column
                    ratio = df[params_value['column_name']][i]/calculate_average(df,params_value['ref'])

                    # write the set id
                    ratio_sheet.cell(row=row_id + i + 1, column=column_id).value = i
                    # write the total time to send and receive request_nb requests
                    ratio_sheet.cell(row=row_id + i + 1, column=column_id + test_count).value = ratio

                    # save the file at each set test to keep track of the test
                    book.save(excel_file)

                test_count += 1
                book.close()

    except openpyxl.utils.exceptions.InvalidFileException:
        print("Error loading the excel file "+excel_file)


def calculate_average(df, column_name):
    '''
    This method calculates and returns the average of the values specified in the column_name
    :param df:data frame returned by pd.read_excel()
    :param column_name: name of column for which to calculate the average values
    :return: average values of the column
    '''
    sum =0
    # loop over the latency values of each column

    for i in range(0, len(df[column_name])):
        sum = sum+df[column_name][i]

    return sum/len(df[column_name])

def get_cpu_overhead (file_path, iterations, wait_time, remote_host=None, remote_user=None, password=None):
    '''
    This method launches the top command on a local or remote host for a certain number of iterations with a wait_time between each iteration
    :param file_path: the file (log.txt) or path to the file (/x/y/log.txt) where the top results should be dumped in the client host (localhost)
    :param iterations: number of iterations that top should perform (-n parameter)
    :param wait_time: number of seconds to wait between each 2 consecutive iterations (-d parameter)
    :param remote_host: ip address of remote host
    :param remote_user: username of remote host
    :param password: password or remote host
    :return:
    '''

    if (remote_host is not None and remote_user is not None):
        #connect to remote server
        remote_session = set_ssh(remote_host, remote_user, password)

        # launch the top command on the remote server
        top_result = remote_session.run('top -d'+str(wait_time)+' -n'+str(iterations)+' -b -uroot')

        #for remote connection dump the top results of the server on the client machine
        file = open(file_path,"w+")
        file.write(str(top_result.stdout))
    else:
        #launch top locally
        os.popen('top -d'+str(wait_time)+' -n'+str(iterations)+' -b -uroot >'+file_path)


def get_RTT (host_address, request_nb):
    '''
    This methods returns the averaged RTT to a certain host of a specific request_nb
    :param host_address: host to ping
    :param request_nb: nb of ICMP request to launch with ping and get the average rtt
    :return: averaged RTT in ms
    '''
    avg_rtt =0
    ping_results = os.popen('ping -c '+ str(request_nb)+" "+ host_address)

    # loop over each line in top results
    for line in ping_results:
        print (line)
        # split each line
        words = line.split()

        if (len(words) is not 0  and words[0] == 'rtt'):
            # get the line corresponding avg rtt
            avg_rtt = words[3].split('/')[1]
            # return the avg RTT in ms
            return avg_rtt
    # return 0 if rtt failed
    return avg_rtt



def launch_requests_client (client, request_nb, mtype, payload_params, total_time):
    '''
    This method keeps launching a request_nb per second for a total of total_time sec
    :param client: the  lurk client object
    :param request_nb: nb of requests to launch per second
    :param mtype: type : rsa, rsa_extended', ecdhe
    :param payload_params: payload parameters as generated by get_payload_params()
    :param total_time: total time of test during which the client keeps on sending requests
    :return:
    '''
    start_time = time()
    ellapsed_time =0
    while (ellapsed_time<total_time):

        start_sec_time = time()
        #send request_nb per sec
        for i in range(0, request_nb):
            client.resolve([{'designation': 'tls12', \
                             'version': 'v1', 'status': "request", \
                             'type': mtype, 'payload': payload_params}])
        #get the time remaining to reach one second
        remaining_sec_time = 1-(time()-start_sec_time)

        if (remaining_sec_time<0):
            print("Launching %s requests exceeded 1 sec"%request_nb)

        #waiting to have the once sec passed
        sleep(remaining_sec_time)

        ellapsed_time = time() - start_time

def launch_requests (total_requests_persec, requests_per_client, total_time, mtype, payload_params, connectivity_conf, clients_pid):
    '''
    This method launch a certain number of clients as processes.
    Each process/client launches a requests_per_client requests per sec for a total of total_time seconds
    The number of clients to launch is calculated as total_requests_persec/requests_per_client
    :param total_requests_persec: Total requests per second to be launched by all the clients
    :param requests_per_client: nb of requests that should be launcheed by each client each second (the time to resolve this number of requests should be <=1sec)
    :param total_time: Total time of the test during which each client should keep sending requests
    :param mtype: type : rsa, rsa_extended', ecdhe
    :param payload_params: payload parameters as generated by get_payload_params()
    :param connectivity_conf: connectivity dictionnary with all the client connectivity information
    :param clients_pid empty list that will hold the launched process ids (this is needed as parameter in case launch requests is started as a process)
    :return: list of client process ids
    '''
    #specify the total number of clients to launch per second
    nb_client = int(total_requests_persec/requests_per_client)

    for i in range(0, nb_client):
        #create the client
        client = set_lurk('client', connectivity_conf=connectivity_conf, resolver_mode='stub')

        #launch a certain nb of requests per client per sec as sub process
        client_process = mp.Process(target=launch_requests_client, args=(client, requests_per_client, mtype, payload_params, total_time), name="client_%d"%i, daemon=True)  ###
        client_process.start()

        #save the clients pid
        clients_pid.append(client_process.pid)

    return clients_pid

def cpu_overhead_test (payload_params, connectivity_conf, file_path, total_requests_persec, requests_per_client,  iterations, wait_time, thread = False,remote_connection = False):
    '''
        This method performs the cpu overhead test on the client and server side.
        It prints the results into the specified payload_params[column_name]+"_client" for client side cpu and payload_params[column_name]+"_server" for server side cpu overhead and place them
        in a client and server folder in the directory specified in file_path
        total_time =(iterations+2)*wait_time: total time of each test in payload_params. After this time the client and server processes are killed
        It prints the parameters in an excel file
    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform.
        each key in payload param should correspond to a key in connectivity conf.
        the other key, values in the dictionary of each item in payload params correspond to those desired for the test and should be similar to those defined in core/conf.py
        2 other keys, values are necessary:
            - column_name: name of the test as we want it to appear in the excel sheet
            - ref: corresponds to a column_name in the payload_params that we want to use to in the ration calculation (latency value of column_name/ latency value of ref)
    For example:
          payload_params = {
            'udp':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param connectivity_conf: dictionary with connectivity information of the client and the server identified by their keys.
        For each server configuration, we should have a client configuration with the same key. This key is also used in payload_params
         For example:
              connectivity_conf= {
              'udp':{
                'type': "udp",
                'ip_address': "127.0.0.1",
                'port': 6789,
                'key': join(data_dir, 'key_tls12_rsa_server.key'),
                'cert': join(data_dir, 'cert_tls12_rsa_server.crt'),
                'key_peer': join(data_dir, 'key_tls12_rsa_client.key'),
                'cert_peer': join(data_dir, 'cert_tls12_rsa_client.crt')
                 'remote_user':remote_user, #user name to set up ssh connection to remote server
                'password': server_password, #password of remote server
                'path_to_erilurk':"Desktop/HyameServer/projects/erilurk"
            }

    }
    :param file_path: path to place te file with the top results (directory)
    :param total_requests_persec:total requests to be sent per sec by all the clients
    :param requests_per_client: number of requests that a client should send per second. This includes the resolve time+waiting time to reach 1 sec
    :param iterations: number of iterations that the top command should performs
    :param wait_time: time to wait between top command iterations
    :param thread: true if multi threading should be used
    :param remote_connection: true if remote connection to server should be performed
    :return:
    '''

    test_params = {
        'connectivity_conf': connectivity_conf,
        'payload_params': payload_params,
        'total_requests_persec': total_requests_persec,
        'requests_per_client':requests_per_client,
        'top_iterations':iterations+2,
        'top_wait_time_btw_ietrations':wait_time,
        'thread': thread,

    }

     # print the parameters to the excel_file
    write_to_excel(file_path+"cpu_overhead.xlsx", "test_params", 1, 1, test_params=test_params)

    for connectivity_key, test_params in payload_params.items():

        #enable remote connection only if remote_user is set
        if remote_connection and 'remote_user' in connectivity_conf[connectivity_key].keys():
            remote = True
        elif remote_connection:
            print("Remote user not provided to enable remote connection --- running locally")
            remote = False
        else:
            remote = False

        #create a client and server folder to hold the cpu overhead results for the client and server side respectively
        client_path = file_path + "client"
        server_path = file_path + "server"

        if os.path.exists(client_path) == False:
            os.makedirs(client_path)

        if os.path.exists(server_path) == False:
            os.makedirs(server_path)

        for params_value in test_params:

            # start server corresponding to the key for each params value to make sure of having accurate results
            server_process_id = start_server(connectivity_conf=connectivity_conf[connectivity_key], thread=thread, remote_connection=remote)


            # generate payload params
            generated_payload_params = get_payload_params(params_value['type'], args=params_value)


            #calculate the total time of the test during which each client will keep launching requests.
            # Total Time can be calculated based on the number of iterations and the wait time between 2 iterations. we add 2 iterations to disregard the first 2 values and
            #have the results averaged over iterations number
            total_time = (iterations+2)*wait_time

            client_file_path = client_path + "/"+params_value['column_name']+"_client.txt"
            server_file_path = server_path + "/"+params_value['column_name']+"_server.txt"

            #save the list of client process ids launching the requests
            client_processes = []

            launch_clients_proc  = mp.Process(target=launch_requests, args = (total_requests_persec, requests_per_client, total_time, params_value['type'], generated_payload_params, connectivity_conf[connectivity_key], client_processes))
            launch_clients_proc.start()

            # run the top command locally
            client_top_process = mp.Process(target=get_cpu_overhead, args = (client_file_path, iterations, wait_time), daemon = True)
            client_top_process.start()

            server_top_process = None
            if (remote == True):
               server_top_process =  mp.Process(target=get_cpu_overhead, args = (server_file_path, iterations, wait_time, connectivity_conf[connectivity_key]['ip_address'],
                                                  connectivity_conf[connectivity_key]['remote_user'],
                                                  connectivity_conf[connectivity_key]['password']), daemon = True)
               server_top_process.start()

               # wait until cient top finish execution before killing the processes and the  server
               server_top_process.join()

            #wait until cient top finish execution before killing the processes and the  server
            client_top_process.join()

             # #kill client processes (even before recieving response) once all top results has been collected
            if (not client_top_process.is_alive() and server_top_process is not None and not server_top_process.is_alive() ):

                #kill client processes
                for process in client_processes:
                    os.kill(process, signal.SIGKILL)


            #stop the server
            if (remote == True):
                stop_server(server_process_id, remote_host=connectivity_conf[connectivity_key]['ip_address'],
                            remote_user=connectivity_conf[connectivity_key]['remote_user'],
                            password=connectivity_conf[connectivity_key]['password'])
            else:
                stop_server(server_process_id)