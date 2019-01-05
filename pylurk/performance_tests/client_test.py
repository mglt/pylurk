import os.path
from  os.path import join
import threading
import pkg_resources
from pylurk.core.conf import default_conf
from pylurk.core.lurk import LurkConf, ConfError, \
                             LurkUDPClient, LurkUDPServer, \
                             LurkTCPClient, LurkTCPServer, \
                             LurkHTTPClient, LurkHTTPServer
from pylurk.extensions.tls12 import LurkExt
from copy import deepcopy
from time import time
import openpyxl
import pandas as pd
from pylurk.performance_tests.server_test import start_server
from pylurk.utils.data_plot import boxplot
from pylurk.utils.excel import write_to_excel


def get_payload_params(mtype, **kwargs):
    '''
    This method will initiate a RequestPayload object based on the mtype and build the payload parameters accordingly given a set of parameters **kwargs
    :param mtype: request type [rsa_master, rsa_master_with_poh, rsa_extended_master, rsa_extended_master_with_poh, ecdhe]
    :param kwargs: parameters needed for building the payload
    :return: payload parameters to be used
    '''

    #get a RequestPayload object based on the mtype
    payload_request = LurkExt('client').get_ext_class()['request', mtype]

    #build payload based on a defined set of paramters
    payload_params = payload_request.build_payload(**kwargs)

    return payload_params

def get_client ( connectivity_conf):
    '''
    This method will initiate and return a Lurk client object based on the connectivity type
    :param conectivity_conf: dictionary for the connectivity information, mainly: type, ip address, port, tls certifications and keys
    :return: LurKClient object corresponding to the protocol
    '''

    clt_conf = LurkConf(deepcopy(default_conf))
    clt_conf.set_role('client')
    clt_conf.set_connectivity(type=connectivity_conf['type'], ip_address=connectivity_conf['ip_address'], port=connectivity_conf['port'])

    connection_type = connectivity_conf['type']

    if connection_type in ['udp', 'udp+dtls']:
        client = LurkUDPClient(conf=clt_conf.get_conf())
    elif connection_type in ['tcp', 'tcp+tls']:
        client = LurkTCPClient(conf=clt_conf.get_conf())
    elif connection_type in ['http', 'https']:
        client = LurkHTTPClient(conf=clt_conf.get_conf())

    return client


def latency_test (payload_params, connectivity_conf, graph_params, sheet_name, fig_path, excel_file = "results.xlsx",  request_nb = 50, set_nb =20):
    '''

    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform. For example:
          payload_params = {'rsa_master_ref_prf_sha256_pfs_sha256':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param connectivity_conf: dictionary with connectivity info. For example:
               connectivity_conf= {
                'type': "udp",  # "udp", "local",
                'ip_address': "127.0.0.1",
                'port': 6789,
               'key' : join( data_dir, 'key_tls12_rsa_server.key'),
              'cert' : join( data_dir, 'cert_tls12_rsa_server.crt'),
              'key_peer' : join( data_dir, 'key_tls12_rsa_client.key'),
              'cert_peer' : join( data_dir, 'cert_tls12_rsa_client.crt')
            }
    :param graph_params a dictionary with all the information of the graph and data to plot (see boxplot() for more info)
    :param sheet_name: name of the excel sheet where the data should be saved
    :param fig_path: path or name of the fig  to save the graph (e.g., results/fig.png)
    :param excel_file:  path of excel file to save the results. The file is created if it does not exist
    :param request_nb: nb of requests  per set to report their latency
    :param set_nb: nb of sets to test
    :return: row_id in the excel sheet where we can start writing without overwriting any existing values
    '''
    #start by setting up a dictionnary of all parameters of the test to write in excel
    test_params = {
        'connectivity_conf': connectivity_conf,
        'payload_params' : payload_params,
        'set_nb':set_nb,
        'request_nb':request_nb,
        'other_info': {'test_type': 'latency', #specify that we are testing the latency
                       'latency_unit': 'seconds', #specify that the reported latency results are in sec
                       }
    }

    parameters_sheet_name = sheet_name+"_params"
    latency_sheet_name = sheet_name+"_Val"
    ratio_sheet_name = sheet_name+"_Ratio"

    # print the parameters to the excel_file
    write_to_excel(excel_file, parameters_sheet_name, 1, 1, test_params = test_params)

    #get latency by running the tests
    run_latency_test(payload_params, connectivity_conf, latency_sheet_name, excel_file = excel_file, request_nb= request_nb, set_nb=set_nb)

    #read the sheet containing the latencies and compute and write the ratios based on the reference results
    calculate_ratio(payload_params, latency_sheet_name, ratio_sheet_name, excel_file = excel_file)

    #create and save graph
    boxplot(ratio_sheet_name, excel_file, graph_params, fig_path)

def run_latency_test (payload_params, connectivity_conf, sheet_name, excel_file="results.xlsx", request_nb=50, set_nb=20):
    '''
    This method performs the latency test and print the results into the specified sheet name
    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform. For example:
          payload_params = {'rsa_master_ref_prf_sha256_pfs_sha256':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param connectivity_conf: dictionary with connectivity info. For example:
               connectivity_conf= {
                'type': "udp",  # "udp", "local",
                'ip_address': "127.0.0.1",
                'port': 6789,
                'key' : join( data_dir, 'key_tls12_rsa_server.key'),
                  'cert' : join( data_dir, 'cert_tls12_rsa_server.crt'),
                  'key_peer' : join( data_dir, 'key_tls12_rsa_client.key'),
                  'cert_peer' : join( data_dir, 'cert_tls12_rsa_client.crt')
            }
    :param sheet_name: name of the excel sheet where the data should be saved
    :param excel_file:  path of excel file to save the results. The file is created if it does not exist
    :param request_nb: nb of requests  per set to report their latency. it corresponds to the number of clients to create and each client generates one request
    :param set_nb: nb of sets to test
    :return: row_id in the excel sheet where we can start writing without overwriting any existing values
    '''

    row_id = 1
    column_id = 1

    # check if excel file exists and create it otherwise
    if (os.path.isfile(excel_file) == False):
        book = openpyxl.Workbook()
        book.save(excel_file)

    book = openpyxl.load_workbook(excel_file)

    # create the sheet if it does not exist in the mentioned excel_file
    if (sheet_name not in book.sheetnames):
        book.create_sheet(sheet_name)
        book.save(excel_file)

    # get the sheet
    sheet = book[sheet_name]

    sheet.cell(row=row_id, column=column_id).value = 'Set'

    test_count = 1

    for key, params_value in payload_params.items():

        # write the column name reflecting the test we are performing
        sheet.cell(row=row_id, column=column_id + test_count).value = params_value['column_name']

        # generate payload params
        generated_payload_params = get_payload_params(params_value['type'], args=params_value)

        # loop over the set nb
        for j in range(0, set_nb):

            # get total processing time for all requests
            time_start = time()

            for i in range(0, request_nb):
                # create a client per request
                client = get_client(connectivity_conf)
                client.resolve([{'designation': 'tls12', \
                                 'version': 'v1', 'status': "request", \
                                 'type': params_value['type'], 'payload': generated_payload_params}])
            # get total processing time for all requests
            time_stop = time()

            # calculate the processing time for requestsNb requests and return it
            total_time = time_stop - time_start

            # write the set id
            sheet.cell(row=row_id + j + 1, column=column_id).value = j
            # write the total time to send and receive request_nb requests
            sheet.cell(row=row_id + j + 1, column=column_id + test_count).value = total_time  # to change columns

            # save the file at each set test to keep track of the test
            book.save(excel_file)

            # print some log info
            print("%s %s resolutions in %s sec." % (set, params_value['column_name'], total_time))

        test_count = test_count + 1
        book.close()

    # row id  is the row nb depicting the title of tests; thus we return an empty row where we can start writing to the file after leaving 2 empty rows
    return row_id + set_nb + 2


def calculate_ratio (payload_params, values_sheet_name, sheet_name, excel_file = "results.xlsx"):
    '''
    This method calculates the ratio values based on the reference values mentioned in payload_params and save them in the mentioned sheet_name
    :param payload_params: dictionary with payload parameters containing the list of tests setups to perform. For example:
          payload_params = {'rsa_master_ref_prf_sha256_pfs_sha256':{
             'type': 'rsa_master',
             'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
             'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
             'prf_hash': 'sha256',
             'freshness_funct': 'sha256',}

        }
    :param values_sheet_name: the sheet name in excel_file containing the values to which we want to calculate the ratio
    :param sheet_name: sheet_name to save the calculates ratio. This sheet is created if it does not exists
    :param excel_file: path to the excel file containing the values to calculate the ratio and to save the latter
    '''
    try:

        book = openpyxl.load_workbook(excel_file)

        # create the results sheet if it does not exist in the mentioned excel_file
        if (sheet_name not in book.sheetnames):
            book.create_sheet(sheet_name)
            book.save(excel_file)


        # get the sheet where we want to store the ratio values
        ratio_sheet = book[sheet_name]

        row_id =1
        column_id =1
        test_count =1

        ratio_sheet.cell(row=row_id, column=column_id).value = 'Set'

        #read the sheet containing the values to which we want to calculate the ratio
        df = pd.read_excel(excel_file, sheet_name=values_sheet_name, engine =None)

        #loop over the different tests (columns)
        for key, params_value in payload_params.items():

            # write the column name reflecting the test
            ratio_sheet.cell(row=row_id, column=column_id + test_count).value = params_value['column_name']

            #loop over the latency values of each column
            for i in range (0, len(df[params_value['column_name']])):

                #calculate the ratio based on the ratio column and print it in the ratio sheet
                ratio = df[params_value['column_name']][i]/df[params_value['ref']][i]

                # write the set id
                ratio_sheet.cell(row=row_id + i + 1, column=column_id).value = i
                # write the total time to send and receive request_nb requests
                ratio_sheet.cell(row=row_id + i + 1, column=column_id + test_count).value = ratio

                # save the file at each set test to keep track of the test
                book.save(excel_file)

            test_count += 1
            book.close()

    except openpyxl.utils.exceptions.InvalidFileException:
        print("Error loading the excel file "+excel_file)




def authentication_methods ():
    payload_params = {'rsa_master_ref_prf_sha256_pfs_sha256': {
        'type': 'rsa_master',
        'column_name': 'rsa_master_ref_prf_sha256_pfs_sha256', # name of column as it appears in excel file, if it contains 'ref', this means that it will be used as ref values
        'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',  # ref column name when calculating ref values
        'prf_hash': 'sha256',
        'freshness_funct': 'sha256',

        },
        'rsa_master_prf_sha384_pfs_sha256': {
            'type': 'rsa_master',
            'column_name': 'rsa_master_prf_sha384_pfs_sha256',
            'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',
            'prf_hash': 'sha384',
            'freshness_funct': 'sha256'
        },
        'rsa_master_prf_sha512_pfs_sha256': {
            'type': "rsa_master",
            'column_name': 'rsa_master_prf_sha512_pfs_sha256',
            'ref': 'rsa_master_ref_prf_sha256_pfs_sha256',
            'prf_hash': 'sha512',
            'freshness_funct': 'sha256'
        },

        'rsa_extended_master_ref_prf_sha256_pfs_sha256': {
            'type': 'rsa_extended_master',
            'column_name': 'rsa_extended_master_ref_prf_sha256_pfs_sha256',
            'ref': 'rsa_extended_master_ref_prf_sha256_pfs_sha256',
            'prf_hash': 'sha256',
            'freshness_funct': 'sha256'

        },
        'rsa_extended_master_prf_sha384_pfs_sha256': {
            'type': "rsa_extended_master",
            'column_name': 'rsa_extended_master_prf_sha384_pfs_sha256',
            'ref': 'rsa_extended_master_ref_prf_sha256_pfs_sha256',
            'prf_hash': "sha384",
            'freshness_funct': "sha256"
        },
        'rsa_extended_master_prf_sha512_pfs_sha256': {
            'type': "rsa_extended_master",
            'column_name': 'rsa_extended_master_prf_sha512_pfs_sha256',
            'ref': 'rsa_extended_master_ref_prf_sha256_pfs_sha256',
            'prf_hash': 'sha512',
            'freshness_funct': 'sha256'
        },

        'ecdhe_ref_sig_sha256rsa_pfs_sha256': {
            'type': 'ecdhe',
            'column_name': 'ecdhe_ref_sig_sha256rsa_pfs_sha256',
            'ref': 'ecdhe_ref_sig_sha256rsa_pfs_sha256',
            'sig_and_hash': ('sha256', 'rsa'),
            'freshness_funct': 'sha256'
        },
        'ecdhe_sig_sha512rsa_pfs_sha256': {
            'type': 'ecdhe',
            'column_name': 'ecdhe_sig_sha512rsa_pfs_sha256',
            'ref': 'ecdhe_ref_sig_sha256rsa_pfs_sha256',
            'sig_and_hash': ('sha512', 'rsa'),
            'freshness_funct': 'sha256'
        },
        # 'ecdhe_sig_sha512rsa_pfs_sha256': {
        #     'type': 'ecdhe',
        #     'column_name': 'ecdhe_sig_sha512rsa_pfs_sha256',
        #     'ref': 'ecdhe_ref_sig_sha256rsa_pfs_sha256',
        #     'sig_and_hash': ('sha512', 'rsa'),
        #     'freshness_funct': 'sha256'
        # },
    }

    data_dir = pkg_resources.resource_filename(__name__, '../data/')
    connectivity_conf= {
        'type': "udp",  # "udp", "local",
        'ip_address': "127.0.0.1",
        'port': 6789,
        'key': join(data_dir, 'key_tls12_rsa_server.key'),
        'cert': join(data_dir, 'cert_tls12_rsa_server.crt'),
        'key_peer': join(data_dir, 'key_tls12_rsa_client.key'),
        'cert_peer': join(data_dir, 'cert_tls12_rsa_client.crt')
    }

    graph_params = {'title': 'Authentication Methods',
                    'xlabel': 'Authentication Methods',
                    'ylabel': 'Latency (sec)',
                    'box_width': 0.5,  # width of each box in the graph
                    'start_position': 1,  # the position of the first box to draw
                    'show_grid': True,  # show grid in the graph
                    'legend': {
                        'location': 'lower right',
                    # location of the legend. Can take one of the following values:'best','upper right','upper left','lower left','lower right','right','center left','center right','lower center','upper center','center'
                        'font_properties': {
                            # 'fontname':'Calibri',
                            'size': '12',
                            # 'weight': 'bold',
                        }
                    },
                    'font_properties': {  # font properties of title, ylabel and xlabel
                        # 'fontname':'Calibri',
                        'size': '14',
                        'weight': 'bold',
                    },
                    'ticks_font_properties': {
                        # 'fontname':'Calibri',
                        'size': '12',
                        # 'weight': 'bold',
                    },
                    # data to plot grouped into multiple group. if no group is desired, a dictionary for each data to plot should be added
                    'groups': [
                        {'tick_label': 'RSA',  # label on xaxis depicting all the data in data
                         'color': ['blue', 'green'],
                         # color of the box of each data in data, set 'White if no color is desired
                         'hatch': ['/', 'o', '*'],
                         # pattern of each box in data. Set '' if no hatch is desired. It can take one of the following patterns = ('-', '+', 'x', '\\', '*', 'o', 'O', '.', '/')
                         'data': ['rsa_master_prf_sha384_pfs_sha256', 'rsa_master_prf_sha512_pfs_sha256'],
                         # colummn name of the data to plot as defined in excel sheet
                         'legends': ['(prf = sha384, pfs = sha256)', '(prf = sha512, pfs = sha256)']
                         # legend corresponding to each data, set None if no legend to be added to a specified data or provide an empty list
                         },
                        {'tick_label': 'RSA_Extended',
                         'color': ['blue', 'green'],  # same color and hatch as previous group to have same legend
                         'hatch': ['/', 'o', '*'],
                         'data': ['rsa_extended_master_prf_sha384_pfs_sha256',
                                  'rsa_extended_master_prf_sha512_pfs_sha256'],
                         'legends': [],  # empty list to have one legend per color as specified in previous group
                         },
                        {'tick_label': 'ECDHE',
                         'color': ['red'],
                         'hatch': ['+'],
                         'data': ['ecdhe_sig_sha512rsa_pfs_sha256'],
                         'legends': ['sig_and_hash = (sha256, rsa)'],
                         },

                    ]}
    start_server(connectivity_conf)
    sheet_name = 'Authentification_latency'
    latency_test (payload_params, connectivity_conf, graph_params, sheet_name, "test.png", excel_file = "results.xlsx", request_nb = 2, set_nb =2)



authentication_methods ()


