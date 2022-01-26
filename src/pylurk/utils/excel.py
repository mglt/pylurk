import openpyxl
import os.path

def print_dict(dictionary, sheet, row_id, column_id):
    '''
    Method that recursively prints nested dictionaries in an excel sheet.
    IMPORTANT NOTE: this method can not be used by itself as it does not save the sheet. It is a helper function that is called within write_to_excel
    :param dictionary: nested dictionary to print (tuples within the dictionary are supported, list of dictionaries within the dictionary is also supported)
    :param sheet: excel sheet object
    :param row_id: row where to start printing
    :param column_id: column where to start printing
    :return: row id after the last used row (row in which new information can be written)
    '''

    for key, value in dictionary.items():
        if isinstance(value, dict):
            sheet.cell(row=row_id, column=column_id).value = key
            row_id = print_dict(value, sheet, row_id, column_id+1)

        else:
            sheet.cell(row=row_id, column=column_id).value = key
            if isinstance(value, tuple):
                v = "( "
                for t in value:
                    v +=str(t)+", "
                v +=" )"
                sheet.cell(row=row_id, column=column_id + 1).value = v
            elif isinstance(value, list):
                v="["

                for t in value:
                    #if list of dictionaries
                    if isinstance(t, dict):
                        row_id = print_dict(t, sheet, row_id+1, column_id + 1)
                    else:

                        v+=str(t)+", "
                        sheet.cell(row=row_id, column=column_id + 1).value = v
            else:
                sheet.cell(row=row_id, column=column_id + 1).value = value
            row_id+=1
    return row_id

def write_to_excel ( excel_file, sheet_name, row_id, column_id, **kwargs):
    '''
    This method loads the excel file and and writes the test parameters in the specified sheet
    :param excel_file: excel file path to write into (this file is created if it does not exist)
    :param sheet_name: the sheet name in the specified excel_file to write into. The sheet is created if it does not exist
    :param row_id: row number where to start writing in the excel_file (takes any value >= 1)
    :param column_id: column where to strat writing in the excel file (takes any value >=  1)
    :param kwargs: the list to write in the specified sheet. This can be a nested dictionary. For example:
            'connectivity_conf': connectivity_conf, #connectivity info including type, ip, port, tls keys and tls certificates
            'secureTLSConnection':secureTLSConnection,
            'payload_params' : payload_params, #all payload parameters including specific lurk tests
            'set_nb':set_nb, #nb of tests to run
            'request_nb':request_nb, #nb of requests to test
            'other_info': {'test_type': 'latency', #specifies the test type [latency, CPU']
                           'latency_unit': 'seconds', #specifies that the reported latency results are in sec
                           'multiThreading': True' #specifies if we are using multithreading
                           'thread_nb': 4 #nb of threads used in the test
                           #any other values needed to be printed can be included here

    }
    :return: the row_id at which we can start writing in the specified sheet (leaves 1 empty row)
    '''

    #check if excel file exists and create it otherwise
    if (os.path.isfile(excel_file) == False):
        book = openpyxl.Workbook()
        book.save(excel_file)

    # load the excel file to write into
    book = openpyxl.load_workbook(excel_file)


    #create the sheet if it does not exist in the mentioned excel_file
    if (sheet_name not in book.sheetnames):
        book.create_sheet(sheet_name)
        book.save(excel_file)

    # load the sheet
    sheet = book[sheet_name]

    #print all the kwargs in the sheet
    row_id = print_dict(kwargs, sheet, row_id, column_id)

    #add an empty row
    row_id += 1
    book.save(excel_file)

    #close the book to prevent any errors in re-loading it
    book.close()

    #return the row where we can start writing other information (results)
    return row_id
